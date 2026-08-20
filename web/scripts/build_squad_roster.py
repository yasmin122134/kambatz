"""Build src/data/squad-roster.json from Excel + platoon-d-roster.json."""
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"c:\Users\yasmi\Downloads\דוק פלוגה ד׳.xlsx")
ROSTER = ROOT / "src/data/platoon-d-roster.json"
OUT = ROOT / "src/data/squad-roster.json"


def parse_xlsx():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    current_squad = None
    squad_headers: list[str] = []
    people = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        col0 = str(row[0]).strip() if row[0] else ""
        if col0.startswith("צוות"):
            current_squad = col0
            if current_squad not in squad_headers:
                squad_headers.append(current_squad)
        first = str(row[1]).strip() if row[1] else ""
        last = str(row[2]).strip() if row[2] else ""
        email = str(row[7]).strip().lower() if row[7] else ""
        if not first or not last or not email:
            continue
        squad_num = squad_headers.index(current_squad) + 1 if current_squad else None
        people.append(
            {
                "name": f"{first} {last}".strip(),
                "email": email,
                "squad": squad_num,
                "squad_label": current_squad,
            }
        )
    return squad_headers, people


def main():
    squad_headers, xlsx_people = parse_xlsx()
    roster = json.loads(ROSTER.read_text(encoding="utf-8"))
    by_email = {r["email"].lower(): r for r in roster}

    matched = []
    unmatched = []
    for p in xlsx_people:
        r = by_email.get(p["email"])
        if not r:
            unmatched.append(p)
            continue
        matched.append(
            {
                "name": r["name"],
                "db_name": r.get("db_name"),
                "email": p["email"],
                "squad": p["squad"],
                "squad_label": p["squad_label"],
            }
        )

    OUT.write_text(
        json.dumps(
            {
                "squads": squad_headers,
                "people": matched,
                "unmatched": unmatched,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"wrote {len(matched)} matched, {len(unmatched)} unmatched -> {OUT}")


if __name__ == "__main__":
    main()
